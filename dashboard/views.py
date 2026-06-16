from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.db.models import Count,Sum, Q
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from registration.models import Student
from course.models import Course, Centre
from datetime import datetime, timedelta
import json
import pandas as pd
from io import BytesIO
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from public.models import Announcement, CarouselImage
from .forms import AnnouncementForm, CarouselImageForm
from datetime import datetime, timedelta, date
import pandas as pd
from django.template.loader import get_template
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.platypus import ( SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer )
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet



def is_admin(user):
    return user.is_authenticated and user.is_staff

@user_passes_test(is_admin)
def dashboard_index(request):
    total_students = Student.objects.count()
    total_courses = Course.objects.count()
    total_centres = Centre.objects.count()
    approved_students = Student.objects.filter(is_approved=True).count()
    pending_approvals = Student.objects.filter(is_approved=False, status='confirmed').count()
    
    # Today's and Yesterday's Registrations
    today = date.today()
    yesterday = today - timedelta(days=1)
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    today_registrations = Student.objects.filter(registration_date__date=today).count()
    yesterday_registrations = Student.objects.filter(registration_date__date=yesterday).count()
    weekly_registrations = Student.objects.filter(registration_date__date__gte=week_ago).count()
    monthly_registrations = Student.objects.filter(registration_date__date__gte=month_ago).count()
    

     # Active Courses with enrollment counts
    active_courses = Course.objects.filter(is_active=True, course_status__in=['open', 'active', 'ongoing']).annotate(
        student_count=Count('students', filter=Q(students__status__in=['pending', 'confirmed']))
    ).order_by('-student_count')[:10]

    # Add fill_rate to each course
    for course in active_courses:
        if course.max_seats and course.max_seats > 0:
            course.fill_rate = min(int((course.student_count / course.max_seats) * 100), 100)
        else:
            course.fill_rate = 0

    # Weekly data for mini chart
    weekly_dates = []
    weekly_counts = []
    for i in range(7, 0, -1):
        day = today - timedelta(days=i)
        count = Student.objects.filter(registration_date__date=day).count()
        weekly_dates.append(day.strftime('%d/%m'))
        weekly_counts.append(count)
    
    # Active Centres with course and student data
    active_centres = []
    for centre in Centre.objects.all():
        courses = centre.available_courses.filter(is_active=True)
        total_students = Student.objects.filter(preferred_centre=centre, status__in=['pending', 'confirmed']).count()
        max_capacity = sum([c.max_seats for c in courses if c.max_seats]) or 1
        occupancy_rate = int((total_students / max_capacity) * 100) if max_capacity else 0
        
        active_centres.append({
            'centre_name': centre.centre_name,
            'centre_contact': centre.centre_contact,
            'courses': courses[:3],  # Limit to 3 courses
            'total_students': total_students,
            'occupancy_rate': min(occupancy_rate, 100)
        })
    
    # Course Mode Statistics
    course_mode_stats = Course.objects.values('mode').annotate(count=Count('id'))
    mode_labels = []
    mode_counts = []
    mode_map = dict(Course.MODE_CHOICES)
    for stat in course_mode_stats:
        mode_labels.append(mode_map.get(stat['mode'], stat['mode']))
        mode_counts.append(stat['count'])
    
    # Registration trends for last 30 days
    dates = []
    registration_counts = []
    for i in range(30):
        day = today - timedelta(days=i)
        count = Student.objects.filter(registration_date__date=day).count()
        dates.append(day.strftime('%d %b'))
        registration_counts.append(count)
    
    dates.reverse()
    registration_counts.reverse()
    
    # Recent registrations
    recent_registrations = Student.objects.all().select_related('course_enrolled', 'preferred_centre')[:10]
    
    # Course enrollment stats for chart
    course_enrollment_stats = Course.objects.annotate(student_count=Count('students')).order_by('-student_count')[:5]
    
    context = {
        'total_students': total_students,
        'total_courses': total_courses,
        'total_centres': total_centres,
        'approved_students': approved_students,
        'pending_approvals': pending_approvals,
        'pending_approvals_count': pending_approvals,
        'recent_registrations': recent_registrations,
        'course_enrollment_stats': course_enrollment_stats,
        'now': datetime.now(),
        'approved_count': approved_students,

        # New data
        'today_registrations': today_registrations,
        'yesterday_registrations': yesterday_registrations,
        'weekly_registrations': weekly_registrations,
        'monthly_registrations': monthly_registrations,
        'active_courses': active_courses,
        'active_courses_count': active_courses.count(),
        'active_centres': active_centres,
        'recent_students_count': monthly_registrations,
        'course_mode_stats': course_mode_stats,
        'mode_labels': mode_labels,
        'mode_counts': mode_counts,
        'dates': dates,
        'registration_counts': registration_counts,
        'weekly_dates': weekly_dates,
        'weekly_counts': weekly_counts,
    }
    return render(request, 'dashboard/index.html', context)




@user_passes_test(is_admin)
def students_list(request):

    students = Student.objects.select_related(
        'course_enrolled',
        'preferred_centre'
    ).all().order_by('-registration_date')

    # ───────────────── Filters ───────────────── #

    course_filter = request.GET.get('course')
    center_filter = request.GET.get('center')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')

    if course_filter:
        students = students.filter(course_enrolled_id=course_filter)

    if center_filter:
        students = students.filter(preferred_centre_id=center_filter)

    if status_filter:
        students = students.filter(status=status_filter)

    if search_query:
        students = students.filter(
            Q(name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(email_id__icontains=search_query) |
            Q(registration_number__icontains=search_query)
        )

    # ───────────────── Export ───────────────── #

    export_type = request.GET.get("export")

    # ================= EXCEL EXPORT ================= #

    if export_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        headers = [
            "Registration No", "Student Name", "Father Name", "Gender", "DOB",
            "Category", "Mobile", "Email", "State", "City", "Institute",
            "Course", "Center", "Status", "Certificate", "Registration Date",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for student in students:
            ws.append([
                student.registration_number,
                student.name,
                student.father_name,
                student.get_gender_display() if student.gender else '',
                student.date_of_birth.strftime('%d-%m-%Y') if student.date_of_birth else '',
                student.get_category_display(),
                student.mobile_number,
                student.email_id,
                student.state or '',
                student.city or '',
                student.institute_name or '',
                student.course_enrolled.course_name if student.course_enrolled else '',
                student.preferred_centre.centre_name if student.preferred_centre else '',
                student.get_status_display(),
                "Issued" if student.is_approved else "Pending",
                student.registration_date.strftime("%d-%m-%Y"),
            ])

        column_widths = {
            "A": 22, "B": 28, "C": 28, "D": 12, "E": 14,
            "F": 12, "G": 16, "H": 35, "I": 20, "J": 18,
            "K": 30, "L": 30, "M": 25, "N": 15, "O": 15, "P": 18,
        }
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
        wb.save(response)
        return response

    # ================= PDF EXPORT ================= #

    if export_type == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="students.pdf"'

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=15, leftMargin=15, topMargin=20, bottomMargin=20
        )

        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph("<b>Students Report</b>", styles['Title']))
        elements.append(Spacer(1, 12))

        data = [[
            "Reg No", "Name", "Gender", "Mobile", "Email",
            "State", "Course", "Center", "Status", "Certificate", "Date",
        ]]

        for student in students:
            data.append([
                student.registration_number,
                student.name,
                student.get_gender_display() if student.gender else '-',
                student.mobile_number,
                student.email_id,
                student.state or '-',
                student.course_enrolled.course_name if student.course_enrolled else '',
                student.preferred_centre.centre_name if student.preferred_centre else '-',
                student.get_status_display(),
                "Issued" if student.is_approved else "Pending",
                student.registration_date.strftime("%d-%m-%Y"),
            ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[80, 90, 50, 65, 110, 60, 95, 75, 55, 55, 60]
        )

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f8fafc")]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))

        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    # ───────────────── Pagination ───────────────── #

    paginator = Paginator(students, 20)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    courses = Course.objects.all()
    centres = Centre.objects.all()

    context = {
        'page_obj': page_obj,
        'courses': courses,
        'centres': centres,
        'selected_course': course_filter,
        'selected_center': center_filter,
        'selected_status': status_filter,
        'search_query': search_query,
    }

    return render(request, 'dashboard/students_list.html', context)



@user_passes_test(is_admin)
def student_detail(request, pk):
    student = get_object_or_404(Student, pk=pk)
    return render(request, 'dashboard/student_detail.html', {'student': student})


@user_passes_test(is_admin)
def approve_certificate(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        if student_id:
            from certificate.models import CertificateDesign, StudentCertificate
            from django.utils import timezone

            try:
                student = Student.objects.get(id=student_id)
                design = CertificateDesign.objects.filter(is_active=True).first()

                if not design:
                    messages.error(request, 'No active certificate design found.')
                    return redirect('dashboard:approve_certificate')

                if student.status == 'confirmed' and not student.is_approved:
                    certificate, created = StudentCertificate.objects.get_or_create(
                        student=student,
                        defaults={
                            'design': design,
                            'issue_date': timezone.now().date(),
                            'issued_by': request.user.get_full_name() or 'NIELIT Administration'
                        }
                    )
                    certificate.save()

                    student.is_approved = True
                    student.status = 'completed'
                    student.save()

                    messages.success(request, f'Certificate approved for {student.name}!')
                else:
                    messages.warning(request, f'{student.name} is not eligible — must be Confirmed and not already approved.')

            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')

        return redirect('dashboard:approve_certificate')

    # GET — show pending approvals
    students = Student.objects.filter(
        status='confirmed',
        is_approved=False
    ).select_related('course_enrolled', 'preferred_centre').order_by('-registration_date')

    context = {
        'students': students,
        'pending_approvals_count': students.count(),
    }
    return render(request, 'dashboard/approve_certificate.html', context)


 
@user_passes_test(is_admin)
def approve_bulk(request):

    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if student_ids:
            from certificate.models import CertificateDesign, StudentCertificate
            from django.utils import timezone
            
            design = CertificateDesign.objects.filter(is_active=True).first()
            
            if not design:
                messages.error(request, 'No active certificate design found. Please create one first.')
                return redirect('dashboard:students_list')
            
            count = 0
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id)
                    if student.status == 'confirmed' and not student.is_approved:
                        # Create certificate (no QR code image saved)
                        certificate, created = StudentCertificate.objects.get_or_create(
                            student=student,
                            defaults={
                                'design': design,
                                'issue_date': timezone.now().date(),
                                'issued_by': request.user.get_full_name() or 'NIELIT Administration'
                            }
                        )
                        
                        # Update student status
                        student.is_approved = True
                        student.status = 'completed'
                        student.save()
                        count += 1
                except Student.DoesNotExist:
                    continue
            
            messages.success(request, f'{count} student(s) have been approved for certification!')
        else:
            messages.warning(request, 'No students selected for approval.')
        return redirect('dashboard:students_list')
    return redirect('dashboard:students_list')


@user_passes_test(is_admin)
def update_status_bulk(request):
    """Bulk update student status"""
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        new_status = request.POST.get('status')
        
        if student_ids and new_status:
            count = 0
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id)
                    student.status = new_status
                    student.save()
                    count += 1
                except Student.DoesNotExist:
                    continue
            messages.success(request, f'{count} student(s) status updated to {new_status}!')
        else:
            messages.warning(request, 'Please select students and a status.')
        return redirect('dashboard:students_list')
    return redirect('dashboard:students_list')

@user_passes_test(is_admin)
def delete_bulk(request):
    """Bulk delete students"""
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if student_ids:
            from certificate.models import StudentCertificate
            # Delete certificates first (due to foreign key)
            StudentCertificate.objects.filter(student_id__in=student_ids).delete()
            # Delete students
            count, _ = Student.objects.filter(id__in=student_ids).delete()
            messages.success(request, f'{count} student(s) have been deleted successfully!')
        else:
            messages.warning(request, 'No students selected for deletion.')
        return redirect('dashboard:students_list')
    return redirect('dashboard:students_list')

@user_passes_test(is_admin)
def update_status(request):

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        new_status = request.POST.get('status')
        try:
            student = Student.objects.get(id=student_id)
            student.status = new_status
            student.save()
            return JsonResponse({'success': True, 'message': f'Status updated to {new_status}!'})
        except Student.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Student not found!'})
    return JsonResponse({'success': False, 'message': 'Invalid request!'})

@user_passes_test(is_admin)
def reports(request):
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    dates = []
    counts = []
    current_date = start_date
    while current_date <= end_date:
        count = Student.objects.filter(registration_date__date=current_date).count()
        dates.append(current_date.strftime('%Y-%m-%d'))
        counts.append(count)
        current_date += timedelta(days=1)
    
    categories = Student.objects.values('category').annotate(count=Count('id'))
    category_labels = []
    category_data = []
    category_map = dict(Student.CATEGORY_CHOICES)
    for c in categories:
        category_labels.append(category_map.get(c['category'], c['category']))
        category_data.append(c['count'])
    
    course_stats = Course.objects.annotate(student_count=Count('students')).filter(student_count__gt=0)
    course_labels = [course.course_name for course in course_stats]
    course_data = [course.student_count for course in course_stats]
    
    context = {
        'dates': json.dumps(dates),
        'counts': json.dumps(counts),
        'category_labels': json.dumps(category_labels),
        'category_data': json.dumps(category_data),
        'course_labels': json.dumps(course_labels),
        'course_data': json.dumps(course_data),
        'total_students': Student.objects.count(),
        'approved_count': Student.objects.filter(is_approved=True).count(),
        'pending_count': Student.objects.filter(is_approved=False, status='confirmed').count(),
    }
    return render(request, 'dashboard/reports.html', context)

@user_passes_test(is_admin)
def export_students_excel(request):
    """Export students data to Excel"""
    students = Student.objects.select_related('course_enrolled', 'preferred_centre').all().values(
        'registration_number', 'name', 'mobile_number', 'email_id', 'date_of_birth', 
        'category', 'father_name', 'course_enrolled__course_name', 'preferred_centre__centre_name',
        'is_approved', 'status', 'registration_date'
    )
    
    df = pd.DataFrame(list(students))
    df.columns = ['Registration Number', 'Name', 'Mobile Number', 'Email', 'Date of Birth', 
                  'Category', 'Father Name', 'Course Enrolled', 'Preferred Centre', 
                  'Certificate Status', 'Registration Status', 'Registration Date']
    
    df['Certificate Status'] = df['Certificate Status'].apply(lambda x: 'Issued' if x else 'Pending')
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students Report', index=False)
    
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=students_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response

@user_passes_test(is_admin)
def export_students_pdf(request):
    """Export students data to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    students = Student.objects.select_related('course_enrolled', 'preferred_centre').all()[:100]
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=students_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=30)
    
    elements.append(Paragraph("NIELIT Students Registration Report", title_style))
    elements.append(Spacer(1, 20))
    
    data = [['Reg Number', 'Name', 'Mobile', 'Email', 'Course', 'Center', 'Status']]
    for student in students:
        data.append([
            student.registration_number,
            student.name,
            student.mobile_number,
            student.email_id[:20],
            student.course_enrolled.course_name[:25],
            student.preferred_centre.centre_name[:20],
            'Approved' if student.is_approved else 'Pending'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


@user_passes_test(is_admin)
def announcement_list(request):
    """List all announcements with drag-drop ordering"""
    announcements = Announcement.objects.all()
    return render(request, 'dashboard/announcement_list.html', {'announcements': announcements})

@user_passes_test(is_admin)
def announcement_create(request):
    """Create new announcement"""
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save()
            messages.success(request, 'Announcement created successfully!')
            return redirect('dashboard:announcement_list')
    else:
        form = AnnouncementForm()
    
    return render(request, 'dashboard/announcement_form.html', {'form': form, 'title': 'Create Announcement'})

@user_passes_test(is_admin)
def announcement_edit(request, pk):
    """Edit announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, instance=announcement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Announcement updated successfully!')
            return redirect('dashboard:announcement_list')
    else:
        form = AnnouncementForm(instance=announcement)
    
    return render(request, 'dashboard/announcement_form.html', {'form': form, 'title': 'Edit Announcement'})

@user_passes_test(is_admin)
def announcement_delete(request, pk):
    """Delete announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    if request.method == 'POST':
        announcement.delete()
        messages.success(request, 'Announcement deleted successfully!')
        return redirect('dashboard:announcement_list')
    
    return render(request, 'dashboard/announcement_confirm_delete.html', {'object': announcement})

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(is_admin)
def announcement_reorder(request):
    """Reorder announcements via drag and drop"""
    try:
        data = json.loads(request.body)
        for item in data.get('items', []):
            Announcement.objects.filter(id=item['id']).update(order=item['order'])
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== CAROUSEL VIEWS ====================

@user_passes_test(is_admin)
def carousel_list(request):
    """List all carousel images with drag-drop ordering"""
    images = CarouselImage.objects.all()
    return render(request, 'dashboard/carousel_list.html', {'images': images})

@user_passes_test(is_admin)
def carousel_create(request):
    """Add new carousel image"""
    if request.method == 'POST':
        form = CarouselImageForm(request.POST, request.FILES)
        if form.is_valid():
            image = form.save()
            messages.success(request, 'Carousel image added successfully!')
            return redirect('dashboard:carousel_list')
    else:
        form = CarouselImageForm()
    
    return render(request, 'dashboard/carousel_form.html', {'form': form, 'title': 'Add Carousel Image'})

@user_passes_test(is_admin)
def carousel_edit(request, pk):
    """Edit carousel image"""
    image = get_object_or_404(CarouselImage, pk=pk)
    
    if request.method == 'POST':
        form = CarouselImageForm(request.POST, request.FILES, instance=image)
        if form.is_valid():
            form.save()
            messages.success(request, 'Carousel image updated successfully!')
            return redirect('dashboard:carousel_list')
    else:
        form = CarouselImageForm(instance=image)
    
    return render(request, 'dashboard/carousel_form.html', {'form': form, 'title': 'Edit Carousel Image'})

@user_passes_test(is_admin)
def carousel_delete(request, pk):
    """Delete carousel image"""
    image = get_object_or_404(CarouselImage, pk=pk)
    
    if request.method == 'POST':
        image.delete()
        messages.success(request, 'Carousel image deleted successfully!')
        return redirect('dashboard:carousel_list')
    
    return render(request, 'dashboard/carousel_confirm_delete.html', {'object': image})

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(is_admin)
def carousel_reorder(request):
    """Reorder carousel images via drag and drop"""
    try:
        data = json.loads(request.body)
        for item in data.get('items', []):
            CarouselImage.objects.filter(id=item['id']).update(order=item['order'])
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
# Add these new functions to dashboard/views.py

@user_passes_test(is_admin)
def revoke_certificate(request):
    """Revoke certificate for a student (single)"""
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        if student_id:
            from certificate.models import StudentCertificate
            
            try:
                student = Student.objects.get(id=student_id)
                
                if student.is_approved and student.status == 'completed':
                    # Update student status
                    student.is_approved = False
                    student.status = 'confirmed'
                    student.save()
                    
                    # Optional: Delete or mark certificate as revoked
                    # StudentCertificate.objects.filter(student=student).delete()
                    
                    messages.success(request, f'Certificate revoked for {student.name}! Status set back to Confirmed.')
                else:
                    messages.error(request, 'Certificate cannot be revoked for this student.')
            except Student.DoesNotExist:
                messages.error(request, 'Student not found.')
        
        return redirect('dashboard:students_list')
    return redirect('dashboard:students_list')


@user_passes_test(is_admin)
def revoke_bulk(request):
    """Bulk revoke certificates for students"""
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if student_ids:
            count = 0
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id)
                    if student.is_approved and student.status == 'completed':
                        student.is_approved = False
                        student.status = 'confirmed'
                        student.save()
                        count += 1
                except Student.DoesNotExist:
                    continue
            messages.success(request, f'{count} certificate(s) have been revoked!')
        else:
            messages.warning(request, 'No students selected.')
        return redirect('dashboard:students_list')
    return redirect('dashboard:students_list')


@user_passes_test(is_admin)
def edit_student(request, pk):
    """Edit student details"""
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        student.name = request.POST.get('name')
        student.father_name = request.POST.get('father_name')
        student.mobile_number = request.POST.get('mobile_number')
        student.email_id = request.POST.get('email_id')
        student.date_of_birth = request.POST.get('date_of_birth')
        student.gender = request.POST.get('gender') or None
        student.category = request.POST.get('category')
        student.state = request.POST.get('state') or None
        student.city = request.POST.get('city') or None
        student.institute_name = request.POST.get('institute_name') or None

        course_id = request.POST.get('course_enrolled')
        centre_id = request.POST.get('preferred_centre')

        if course_id:
            student.course_enrolled_id = course_id
        if centre_id:
            student.preferred_centre_id = centre_id

        student.save()
        messages.success(request, f'Student {student.name} details updated successfully!')
        return redirect('dashboard:students_list')

    courses = Course.objects.filter(is_active=True)
    centres = Centre.objects.all()

    context = {
        'student': student,
        'courses': courses,
        'centres': centres,
    }
    return render(request, 'dashboard/edit_student_modal.html', context)


@user_passes_test(is_admin)
def get_student_details(request, pk):
    """Get student details for modal view"""
    student = get_object_or_404(Student, pk=pk)
    data = {
        'id': str(student.id),
        'registration_number': student.registration_number,
        'name': student.name,
        'father_name': student.father_name,
        'date_of_birth': student.date_of_birth.strftime('%d %B %Y'),
        'gender': student.get_gender_display() if student.gender else 'NA',
        'category': student.get_category_display(),
        'mobile_number': student.mobile_number,
        'email_id': student.email_id,
        'state': student.state or 'NA',
        'city': student.city or 'NA',
        'institute_name': student.institute_name or 'NA',
        'course_name': student.course_enrolled.course_name,
        'centre_name': student.preferred_centre.centre_name if student.preferred_centre else 'NA',
        'registration_date': student.registration_date.strftime('%d %B %Y, %I:%M %p'),
        'status': student.get_status_display(),
        'is_approved': student.is_approved,
        'certificate_number': student.student_certificate.certificate_number if hasattr(student, 'student_certificate') else 'Not issued',
    }
    return JsonResponse(data)



from django.views.decorators.http import require_POST
from kyndryl.models import KyndrylRegistration


def kyndryl_student_list(request):
    """Main list view with filters, search, export."""
    qs = KyndrylRegistration.objects.all()

    # ── Filters ──
    selected_qualification = request.GET.get('qualification', '')
    selected_category      = request.GET.get('category', '')
    selected_employment    = request.GET.get('employment', '')
    selected_cloud         = request.GET.get('cloud', '')
    search_query           = request.GET.get('search', '').strip()

    if selected_qualification:
        qs = qs.filter(highest_qualification=selected_qualification)
    if selected_category:
        qs = qs.filter(category=selected_category)
    if selected_employment:
        qs = qs.filter(current_employment_status=selected_employment)
    if selected_cloud:
        qs = qs.filter(expertise_in_cloud_computing=selected_cloud)
    if search_query:
        qs = qs.filter(
            Q(name__icontains=search_query) |
            Q(email_id__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(registration_number__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(state__icontains=search_query)
        )

    # ── Export ──
    export = request.GET.get('export', '')
    if export == 'excel':
        return _export_excel(qs)
    if export == 'pdf':
        return _export_pdf(qs)

    # ── Paginate ──
    paginator  = Paginator(qs, 20)
    page_obj   = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj':               page_obj,
        'search_query':           search_query,
        'selected_qualification': selected_qualification,
        'selected_category':      selected_category,
        'selected_employment':    selected_employment,
        'selected_cloud':         selected_cloud,
        # Choice tuples for filter dropdowns
        'qualification_choices':  KyndrylRegistration.QUALIFICATION_CHOICES,
        'category_choices':       KyndrylRegistration.CATEGORY_CHOICES,
        'employment_choices':     KyndrylRegistration.CURRENT_EMPLOYMENT_STATUS_CHOICES,
        'cloud_choices':          KyndrylRegistration.CLOUD_COMPUTING_CHOICES,
        'beneficiary_choices':    KyndrylRegistration.BENEFICIARY_BELONGING_CHOICES,
    }
    return render(request, 'dashboard/kyndryl_students.html', context)


def kyndryl_student_detail_api(request, pk):
    """AJAX endpoint — returns student JSON for the modal."""
    student = get_object_or_404(KyndrylRegistration, pk=pk)
    beneficiary_display = dict(KyndrylRegistration.BENEFICIARY_BELONGING_CHOICES).get(
        student.beneficiary_belonging, student.beneficiary_belonging or '—'
    )
    data = {
        'id':                               str(student.id),
        'registration_number':              student.registration_number,
        'name':                             student.name,
        'father_name':                      student.father_name or '—',
        'mother_name':                      student.mother_name or '—',
        'father_occupation':                student.father_occupation or '—',
        'mother_occupation':                student.mother_occupation or '—',
        'date_of_birth':                    student.date_of_birth.strftime('%d %b %Y') if student.date_of_birth else '—',
        'gender':                           student.get_gender_display() if student.gender else '—',
        'category':                         student.get_category_display(),
        'mobile_number':                    student.mobile_number,
        'email_id':                         student.email_id,
        'address':                          student.address or '—',
        'city':                             student.city or '—',
        'state':                            student.state or '—',
        'pin_code':                         student.pin_code,
        'aadhar_number':                    student.aadhar_number,
        'highest_qualification':            student.get_highest_qualification_display() if student.highest_qualification else '—',
        'current_employment_status':        student.get_current_employment_status_display() if student.current_employment_status else '—',
        'expertise_in_cloud_computing':     student.get_expertise_in_cloud_computing_display() if student.expertise_in_cloud_computing else '—',
        'beneficiary_belonging':            beneficiary_display,
        'registration_date':                student.registration_date.strftime('%d %b %Y, %I:%M %p'),
        'photo_url':                        student.photo.url if student.photo else None,
    }
    return JsonResponse(data)


def kyndryl_student_edit(request, pk):
    """Edit view — reuses your KyndrylRegistrationForm."""
    from kyndryl.forms import KyndrylRegistrationForm
    student = get_object_or_404(KyndrylRegistration, pk=pk)

    if request.method == 'POST':
        form = KyndrylRegistrationForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f'{student.name} updated successfully.')
            return redirect('dashboard:kyndryl_students')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = KyndrylRegistrationForm(instance=student)

    return render(request, 'dashboard/kyndryl_student_edit.html', {
        'form': form,
        'student': student,
    })


@require_POST
def kyndryl_student_delete_bulk(request):
    """Delete one or many students."""
    ids = request.POST.getlist('student_ids')
    if ids:
        deleted, _ = KyndrylRegistration.objects.filter(id__in=ids).delete()
        messages.success(request, f'{deleted} student(s) deleted.')
    else:
        messages.warning(request, 'No students selected.')
    return redirect('dashboard:kyndryl_students')


@require_POST
def kyndryl_student_delete_single(request, pk):
    student = get_object_or_404(KyndrylRegistration, pk=pk)
    name = student.name
    student.delete()
    messages.success(request, f'{name} deleted successfully.')
    return redirect('dashboard:kyndryl_students')


# ── Export helpers ──────────────────────────────────────────────────────────

def _export_excel(qs):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kyndryl Registrations'

    headers = [
        'Reg. Number', 'Name', 'Father Name', 'Gender', 'Date of Birth',
        'Mobile', 'Email', 'Aadhaar', 'Category', 'Qualification',
        'Employment Status', 'Cloud Expertise', 'Beneficiary',
        'Address', 'City', 'State', 'PIN Code', 'Registration Date',
    ]

    # Header row styling
    header_fill = PatternFill('solid', fgColor='0F172A')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    beneficiary_map = dict(KyndrylRegistration.BENEFICIARY_BELONGING_CHOICES)

    for row_idx, s in enumerate(qs, 2):
        ws.append([
            s.registration_number,
            s.name,
            s.father_name or '',
            s.get_gender_display() if s.gender else '',
            s.date_of_birth.strftime('%d/%m/%Y') if s.date_of_birth else '',
            s.mobile_number,
            s.email_id,
            s.aadhar_number,
            s.get_category_display(),
            s.get_highest_qualification_display() if s.highest_qualification else '',
            s.get_current_employment_status_display() if s.current_employment_status else '',
            s.get_expertise_in_cloud_computing_display() if s.expertise_in_cloud_computing else '',
            beneficiary_map.get(s.beneficiary_belonging, s.beneficiary_belonging or ''),
            s.address or '',
            s.city or '',
            s.state or '',
            s.pin_code,
            s.registration_date.strftime('%d/%m/%Y %H:%M'),
        ])
        # Alternate row shading
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill('solid', fgColor='F1F5F9')

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="kyndryl_registrations.xlsx"'
    wb.save(response)
    return response


def _export_pdf(qs):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return HttpResponse('reportlab not installed. Run: pip install reportlab', status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="kyndryl_registrations.pdf"'

    doc    = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#0F172A'), spaceAfter=4)
    sub_style   = ParagraphStyle('sub',   parent=styles['Normal'],   fontSize=9,  textColor=colors.HexColor('#64748B'), spaceAfter=12)
    story.append(Paragraph('NIELIT × Kyndryl — DevSecOps Registrations', title_style))
    story.append(Paragraph(f'Total records: {qs.count()}', sub_style))
    story.append(Spacer(1, 0.3*cm))

    headers = ['Reg. No.', 'Name', 'Mobile', 'Email', 'Category', 'Qualification', 'City', 'State', 'Reg. Date']
    data    = [headers]

    cell_style = ParagraphStyle('cell', fontSize=7, leading=9)
    for s in qs:
        data.append([
            Paragraph(s.registration_number, cell_style),
            Paragraph(s.name, cell_style),
            Paragraph(s.mobile_number, cell_style),
            Paragraph(s.email_id, cell_style),
            Paragraph(s.get_category_display(), cell_style),
            Paragraph(s.get_highest_qualification_display() if s.highest_qualification else '—', cell_style),
            Paragraph(s.city or '—', cell_style),
            Paragraph(s.state or '—', cell_style),
            Paragraph(s.registration_date.strftime('%d/%m/%Y'), cell_style),
        ])

    col_widths = [3.5*cm, 4*cm, 2.8*cm, 5*cm, 2.2*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#0F172A')),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0),  8),
        ('ALIGN',       (0,0), (-1,0),  'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('FONTSIZE',    (0,1), (-1,-1), 7),
        ('GRID',        (0,0), (-1,-1), 0.4, colors.HexColor('#E2E8F0')),
        ('VALIGN',      (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING',  (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING',(0,0), (-1,-1), 4),
    ]))
    story.append(table)
    doc.build(story)
    return response