"""Export helpers.

Requires:  pip install openpyxl reportlab
"""
from django.http import HttpResponse
from django.utils import timezone


def _filename(form, ext):
    stamp = timezone.now().strftime("%Y%m%d-%H%M")
    safe = form.slug or "form"
    return f"{safe}-responses-{stamp}.{ext}"


def build_rows(form):
    """Return (headers, list_of_row_lists) for a form's responses."""
    fields = list(form.ordered_fields)
    headers = ["#", "Submitted at"] + [f.label for f in fields]
    rows = []
    responses = form.responses.all().prefetch_related("files").order_by("submitted_at")
    for i, resp in enumerate(responses, start=1):
        row = [i, timezone.localtime(resp.submitted_at).strftime("%Y-%m-%d %H:%M")]
        for f in fields:
            row.append(resp.value_for(f))
        rows.append(row)
    return headers, rows


def export_excel(form):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    headers, rows = build_rows(form)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=str(val) if val is not None else "")

    # auto width (rough)
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 50)

    ws.freeze_panes = "A2"

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{_filename(form, "xlsx")}"'
    wb.save(resp)
    return resp


def export_pdf(form):
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    headers, rows = build_rows(form)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=form.title)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(form.title, styles["Title"]),
        Paragraph(f"Total responses: {len(rows)}", styles["Normal"]),
        Spacer(1, 12),
    ]

    # wrap cell text so long values don't overflow
    body = styles["BodyText"]
    body.fontSize = 8
    data = [[Paragraph(str(h), styles["Heading5"]) for h in headers]]
    for row in rows:
        data.append([Paragraph(str(v), body) for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    pdf = buf.getvalue()
    buf.close()
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{_filename(form, "pdf")}"'
    resp.write(pdf)
    return resp